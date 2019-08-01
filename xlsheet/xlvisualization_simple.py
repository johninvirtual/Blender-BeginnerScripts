import bpy  # this module is blender specific
import openpyxl  # importing openpyxl -- can be installed via pip -- should be installed for this code to work

# loading the workbook
wb = openpyxl.load_workbook(filename = r"your_xlsx_filepath_here")  # make sure to add your xlsx file path here
sheet = wb["Sheet1"]

# makeTitle function
def make_title(title, mesh, size):
    """
    This function creates a Title block on every cube object
    """
    txt = title.value  # title value is extracted from colomn 1
    x = mesh[i].location.x  # getting x location of the cube
    y = mesh[i].location.y  # getting y location of the cube
    z = mesh[i].location.z  # getting z location of the cube
    bpy.ops.object.text_add(radius=.7*size, rotation=(1.5708,0,1.5708),location=(x,y-size,z+size))  # adding text
    ob=bpy.context.object
    ob.data.body = txt  # setting text to value from colomn 1


mesh=[]  # initializing a list to store mesh data
name_ext = ""  # initializing a string to store mesh name extension. eg:- cube.001 -- .001 is the extension
ext=0  # increment value in name extension
gap=0  # intializing gap value. increment value inside for loop
i=0  # initializing i to access mesh data


for row in range(1,sheet.max_row+1):
    title = sheet.cell(row,1)  # accessing the value in colomn 1
    cell = sheet.cell(row,2)  # accessing the value in colomn 2
    bpy.ops.mesh.primitive_cube_add(size=int(cell.value), location=(0,gap,0))  # creating cube based on value in colomn 2
    mesh.append(bpy.data.objects["Cube"+name_ext])  # storing mesh data in mesh list
    make_title(title, mesh, cell.value)  # calling makeTitle function 
    i+=1  # incrementing i for next mesh data
    gap+=100  # set the gap value between cube here
    ext+=1  # incrementing name extension
    name_ext = f".00{ext}"  # incrementing and storing name extension
